import sqlite3
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

DB_PATH = "db/nifty100.db"


def load_peer_data():

    conn = sqlite3.connect(DB_PATH)

    df = pd.read_sql("""

    SELECT

        pg.peer_group_name,
        pg.company_id,
        c.company_name AS company_name,
        pg.is_benchmark,

        fr.year,
        fr.return_on_equity_pct,
        c.roce_percentage,
        fr.net_profit_margin_pct,
        fr.debt_to_equity,
        fr.free_cash_flow_cr,
        fr.pat_cagr_5yr,
        fr.revenue_cagr_5yr,
        fr.eps_cagr_5yr,
        fr.interest_coverage,
        fr.asset_turnover,
        fr.composite_quality_score

    FROM peer_groups pg

    LEFT JOIN financial_ratios fr
    ON pg.company_id = fr.company_id

    LEFT JOIN companies c
    ON pg.company_id = c.id

    WHERE fr.year = (

        SELECT MAX(f2.year)

        FROM financial_ratios f2

        WHERE f2.company_id = fr.company_id

    )

    """, conn)

    conn.close()

    return df


def percentile(series):

    return (
        series
        .rank(
            pct=True,
            na_option="keep"
        ) * 100
    )


def calculate_peer_percentiles(df):

    all_groups = []

    metrics = [

        "return_on_equity_pct",
        "roce_percentage",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "pat_cagr_5yr",
        "revenue_cagr_5yr",
        "eps_cagr_5yr",
        "interest_coverage",
        "asset_turnover"

    ]

    for group in df["peer_group_name"].unique():

        group_df = df[
            df["peer_group_name"] == group
        ].copy()

        for metric in metrics:

            if metric == "debt_to_equity":

                group_df["debt_to_equity_percentile"] = (
                    100 - percentile(group_df[metric])
                )

            else:

                group_df[f"{metric}_percentile"] = percentile(
                    group_df[metric]
                )

        all_groups.append(group_df)

    result = pd.concat(
        all_groups,
        ignore_index=True
    )

    return result
def save_peer_percentiles(df):

    conn = sqlite3.connect(DB_PATH)

    export_rows = []

    metrics = [

        "return_on_equity_pct",
        "roce_percentage",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "pat_cagr_5yr",
        "revenue_cagr_5yr",
        "eps_cagr_5yr",
        "interest_coverage",
        "asset_turnover"

    ]

    for _, row in df.iterrows():

        for metric in metrics:

            percentile_column = (
                "debt_to_equity_percentile"
                if metric == "debt_to_equity"
                else f"{metric}_percentile"
            )

            export_rows.append({

                "company_id": row["company_id"],
                "peer_group_name": row["peer_group_name"],
                "metric": metric,
                "value": row[metric],
                "percentile_rank": row[percentile_column],
                "year": row["year"]

            })

    export_df = pd.DataFrame(export_rows)

    export_df.to_sql(
        "peer_percentiles",
        conn,
        if_exists="replace",
        index=False
    )

    conn.close()

    print("✅ peer_percentiles table created successfully")

    return export_df

def export_peer_excel(df):

    output = "output/peer_comparison.xlsx"

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFF2CC",
        end_color="FFF2CC"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    benchmark_fill = PatternFill(
        fill_type="solid",
        start_color="FFD966",
        end_color="FFD966"
    )

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        for group in sorted(df["peer_group_name"].unique()):

            sheet = df[
                df["peer_group_name"] == group
            ].copy()
            sheet = sheet.reindex(columns=[

    "company_id",
    "company_name",
    "is_benchmark",
    "year",

    "return_on_equity_pct",
    "roce_percentage",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
    "composite_quality_score",

    "return_on_equity_pct_percentile",
    "roce_percentage_percentile",
    "net_profit_margin_pct_percentile",
    "debt_to_equity_percentile",
    "free_cash_flow_cr_percentile",
    "pat_cagr_5yr_percentile",
    "revenue_cagr_5yr_percentile",
    "eps_cagr_5yr_percentile",
    "interest_coverage_percentile",
    "asset_turnover_percentile"

])

            sheet.to_excel(
                writer,
                sheet_name=group[:31],
                index=False
            )
            print(f"Generating Sheet : {group}")

            worksheet = writer.sheets[group[:31]]

            worksheet.freeze_panes = "A2"

            # -------------------
            # Header Style
            # -------------------

            for cell in worksheet[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # -------------------
            # Auto Width
            # -------------------

            for column_cells in worksheet.columns:

                max_length = 0

                column = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                worksheet.column_dimensions[
                    column
                ].width = max_length + 3

            # -------------------
            # Benchmark Highlight
            # -------------------

            benchmark_col = None

            for col in range(
                1,
                worksheet.max_column + 1
            ):

                if worksheet.cell(
                    row=1,
                    column=col
                ).value == "is_benchmark":

                    benchmark_col = col
                    break

            if benchmark_col:

                for row in range(
                    2,
                    worksheet.max_row + 1
                ):

                    value = worksheet.cell(
                        row=row,
                        column=benchmark_col
                    ).value

                    if str(value) == "1":

                        for c in range(
                            1,
                            worksheet.max_column + 1
                        ):

                            worksheet.cell(
                                row=row,
                                column=c
                            ).fill = benchmark_fill

            # -------------------
            # Percentile Colours
            # -------------------

            for col in range(
                1,
                worksheet.max_column + 1
            ):

                header = worksheet.cell(
                    row=1,
                    column=col
                ).value

                if (
                    header is not None
                    and "percentile" in str(header)
                ):

                    for row in range(
                        2,
                        worksheet.max_row + 1
                    ):

                        cell = worksheet.cell(
                            row=row,
                            column=col
                        )

                        if cell.value is None:
                            continue

                        try:
                            value = float(cell.value)

                        except:
                            continue

                        if value >= 75:

                            cell.fill = green_fill

                        elif value <= 25:

                            cell.fill = red_fill

                        else:

                            cell.fill = yellow_fill

            # -------------------
            # Median Row
            # -------------------

            median_row = worksheet.max_row + 1

            worksheet.cell(
                row=median_row,
                column=1
            ).value = "Median"
            for cell in worksheet[median_row]:
                cell.font = Font(
                    bold=True
                    )
                cell.fill = header_fill

            for col in range(
                2,
                worksheet.max_column + 1
            ):

                values = []

                for row in range(
                    2,
                    median_row
                ):

                    value = worksheet.cell(
                        row=row,
                        column=col
                    ).value

                    if isinstance(
                        value,
                        (int, float)
                    ):

                        values.append(value)

                if values:

                    worksheet.cell(
                        row=median_row,
                        column=col
                    ).value = round(
                        pd.Series(values).median(),
                        2
                    )

    print("✅ peer_comparison.xlsx generated")

def main():

    print("=" * 60)
    print("Loading Peer Data...")
    print("=" * 60)

    df = load_peer_data()
    print(f"\nLoaded {len(df)} companies.")

    # Companies without peer group
    no_peer = df[df["peer_group_name"].isna()]

    if not no_peer.empty:

        print("\nNo peer group assigned:")

        print(
            no_peer["company_id"].tolist()
        )

    # Remove companies without peer group
    df = df.dropna(
        subset=["peer_group_name"]
    )

    print("\nLoaded Peer Data\n")
    print(df.head())

    # Calculate Percentiles
    df = calculate_peer_percentiles(df)

    # Save into SQLite
    export_df = save_peer_percentiles(df)
    print(f"✅ Rows written to SQLite : {len(export_df)}")


    # ---------------------------------
    # Export Final Excel Report
    # ---------------------------------

    export_peer_excel(df)

    print("\n" + "=" * 60)
    print("Day 20 Summary")
    print("=" * 60)

    print(f"Peer Groups Processed : {df['peer_group_name'].nunique()}")

    print(f"Companies Processed   : {len(df)}")

    print(f"SQLite Rows Written   : {len(export_df)}")

    print("Excel Report          : output/peer_comparison.xlsx")

    print("\nPreview\n")

    print(df.head())

    print("\n✅ Day 20 Completed Successfully!")


if __name__ == "__main__":
    main()