import pandas as pd
import yaml

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import get_column_letter


def create_screener_workbook(results):
    """
    Create screener_output.xlsx

    Features
    --------
    ✓ One worksheet per screener
    ✓ Styled headers
    ✓ Freeze top row
    ✓ Auto filter
    ✓ Auto column width
    ✓ Threshold colour coding
    ✓ Thin borders
    ✓ Number formatting
    ✓ Professional formatting
    """

    print("=" * 60)
    print("Creating screener_output.xlsx")
    print("=" * 60)

    output_path = "output/screener_output.xlsx"

    with open("config/screener_config.yaml", "r") as file:
        config = yaml.safe_load(file)

    # =====================================================
    # Styles
    # =====================================================

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    thin_side = Side(
        style="thin",
        color="D9D9D9"
    )

    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )

    # =====================================================
    # Workbook
    # =====================================================

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:

        for sheet_name, dataframe in results.items():

            dataframe.to_excel(
                writer,
                sheet_name=sheet_name[:31],
                index=False
            )

            worksheet = writer.sheets[sheet_name[:31]]

            # =================================================
            # Freeze Header
            # =================================================

            worksheet.freeze_panes = "A2"

            # =================================================
            # Auto Filter
            # =================================================

            worksheet.auto_filter.ref = worksheet.dimensions

            # =================================================
            # Header Height
            # =================================================

            worksheet.row_dimensions[1].height = 24

            # =================================================
            # Header Formatting
            # =================================================

            for cell in worksheet[1]:

                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # =================================================
            # Borders + Number Formatting
            # =================================================

            for row in worksheet.iter_rows(min_row=2):

                for cell in row:

                    cell.border = border

                    if isinstance(cell.value, float):
                        cell.number_format = "0.00"

            # =================================================
            # Auto Column Width
            # =================================================

            for column_cells in worksheet.columns:

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                max_length = 0

                for cell in column_cells:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                worksheet.column_dimensions[
                    column_letter
                ].width = min(max_length + 3, 35)

            # =================================================
            # Threshold Colour Coding
            # =================================================

            sheet_rules = config.get(sheet_name, {})

            headers = {
                worksheet.cell(
                    row=1,
                    column=col
                ).value: col
                for col in range(
                    1,
                    worksheet.max_column + 1
                )
            }

            for metric, rule in sheet_rules.items():

                if metric not in headers:
                    continue

                column_no = headers[metric]

                for row in range(2, worksheet.max_row + 1):

                    cell = worksheet.cell(
                        row=row,
                        column=column_no
                    )

                    if cell.value is None:
                        continue

                    try:
                        value = float(cell.value)

                    except (TypeError, ValueError):
                        continue

                    passed = True

                    if "min" in rule:
                        passed = value >= rule["min"]

                    if "max" in rule:
                        passed = value <= rule["max"]

                    if passed:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill

    print(f"✅ {output_path} generated successfully")