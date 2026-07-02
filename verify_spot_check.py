"""Generate a Day 12 Excel spot-check workbook for ratio validation."""

from __future__ import annotations

import re
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.properties import CalcProperties


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "db" / "nifty100.db"
OUTPUT_PATH = BASE_DIR / "output" / "spot_check.xlsx"
PREFERRED_COMPANIES = ["ABB", "TCS", "RELIANCE"]
SPOT_CHECK_COLUMNS = [
    "Company",
    "Year",
    "Sales",
    "Net Profit",
    "Equity Capital",
    "Reserves",
    "Database ROE",
    "Manual ROE",
    "Difference (%)",
    "Database Revenue CAGR",
    "Manual Revenue CAGR",
    "Difference (%)",
    "Status",
]

PASS_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def connect_db(db_path: Path) -> sqlite3.Connection:
    """Connect to the SQLite database with row access by column name."""

    connection = sqlite3.connect(db_path)
    connection.row_factory = sqlite3.Row
    return connection


def extract_year_number(year_value: Any) -> int | None:
    """Extract the four-digit year from a stored year label."""

    if year_value is None:
        return None

    match = re.search(r"(\d{4})", str(year_value))
    if match is None:
        return None

    return int(match.group(1))


def safe_float(value: Any) -> float | None:
    """Convert a value to float when possible."""

    if value is None:
        return None

    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def manual_roe(net_profit: float | None, equity_capital: float | None, reserves: float | None) -> float | None:
    """Calculate ROE manually from the raw balance sheet values."""

    if net_profit is None or equity_capital is None or reserves is None:
        return None

    denominator = equity_capital + reserves
    if denominator <= 0:
        return None

    return round((net_profit / denominator) * 100, 2)


def manual_revenue_cagr(current_sales: float | None, historical_sales: float | None) -> float | None:
    """Calculate five-year revenue CAGR manually."""

    if current_sales is None or historical_sales is None:
        return None

    if current_sales <= 0 or historical_sales <= 0:
        return None

    return round(((current_sales / historical_sales) ** (1 / 5) - 1) * 100, 2)


def build_history_map(rows: list[sqlite3.Row]) -> dict[str, dict[int, float | None]]:
    """Build a company/year sales lookup for historical CAGR calculations."""

    history: dict[str, dict[int, float | None]] = defaultdict(dict)
    for row in rows:
        company_id = row["company_id"]
        year_number = extract_year_number(row["year"])
        sales = safe_float(row["sales"])
        if company_id is None or year_number is None:
            continue
        history[company_id][year_number] = sales
    return history


def fetch_candidate_rows(
    connection: sqlite3.Connection,
) -> list[sqlite3.Row]:
    """Fetch rows that already have database revenue CAGR values."""

    query = """
        SELECT
            fr.company_id,
            fr.year,
            pnl.sales,
            pnl.net_profit,
            bs.equity_capital,
            bs.reserves,
            fr.return_on_equity_pct,
            fr.revenue_cagr_5yr
        FROM financial_ratios AS fr
        JOIN profitandloss AS pnl
            ON pnl.company_id = fr.company_id
           AND pnl.year = fr.year
        JOIN balancesheet AS bs
            ON bs.company_id = fr.company_id
           AND bs.year = fr.year
        WHERE fr.company_id IS NOT NULL
          AND fr.year IS NOT NULL
          AND fr.revenue_cagr_5yr IS NOT NULL
          AND fr.return_on_equity_pct IS NOT NULL
          AND pnl.sales IS NOT NULL
          AND pnl.net_profit IS NOT NULL
          AND bs.equity_capital IS NOT NULL
          AND bs.reserves IS NOT NULL
        ORDER BY fr.company_id, fr.year;
    """

    cursor = connection.execute(query)
    return cursor.fetchall()


def choose_spot_check_rows(
    candidates: list[sqlite3.Row],
    sales_history: dict[str, dict[int, float | None]],
) -> list[dict[str, Any]]:
    """Select three valid company rows with priority for known tickers."""

    grouped: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for row in candidates:
        grouped[row["company_id"]].append(row)

    for rows in grouped.values():
        rows.sort(
            key=lambda item: (
                extract_year_number(item["year"]) or -1,
                str(item["year"]),
            ),
            reverse=True,
        )

    selected: list[dict[str, Any]] = []
    used_companies: set[str] = set()

    def add_best_row(company_id: str) -> None:
        if company_id in used_companies:
            return

        for row in grouped.get(company_id, []):
            current_year = extract_year_number(row["year"])
            if current_year is None:
                continue

            historical_sales = sales_history.get(company_id, {}).get(current_year - 5)
            current_sales = safe_float(row["sales"])
            net_profit = safe_float(row["net_profit"])
            equity_capital = safe_float(row["equity_capital"])
            reserves = safe_float(row["reserves"])
            db_roe = safe_float(row["return_on_equity_pct"])
            db_revenue_cagr = safe_float(row["revenue_cagr_5yr"])

            manual_roe_value = manual_roe(net_profit, equity_capital, reserves)
            manual_revenue_cagr_value = manual_revenue_cagr(
                current_sales,
                historical_sales,
            )

            if (
                db_roe is None
                or db_revenue_cagr is None
                or manual_roe_value is None
                or manual_revenue_cagr_value is None
            ):
                continue

            selected.append(
                {
                    "company_id": row["company_id"],
                    "year": row["year"],
                    "sales": current_sales,
                    "net_profit": net_profit,
                    "equity_capital": equity_capital,
                    "reserves": reserves,
                    "database_roe": db_roe,
                    "manual_roe": manual_roe_value,
                    "roe_difference": abs(db_roe - manual_roe_value),
                    "database_revenue_cagr": db_revenue_cagr,
                    "historical_sales": historical_sales,
                    "manual_revenue_cagr": manual_revenue_cagr_value,
                    "revenue_cagr_difference": abs(db_revenue_cagr - manual_revenue_cagr_value),
                }
            )
            used_companies.add(company_id)
            return

    for company_id in PREFERRED_COMPANIES:
        add_best_row(company_id)

    if len(selected) < 3:
        remaining_companies = sorted(
            (company_id for company_id in grouped if company_id not in used_companies),
            key=lambda company_id: (
                max(
                    extract_year_number(row["year"]) or -1
                    for row in grouped[company_id]
                ),
                company_id,
            ),
            reverse=True,
        )
        for company_id in remaining_companies:
            add_best_row(company_id)
            if len(selected) == 3:
                break

    return selected[:3]


def create_workbook(rows: list[dict[str, Any]], output_path: Path) -> None:
    """Write the selected spot-check rows to an Excel workbook."""

    workbook = Workbook()
    workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    worksheet = workbook.active
    worksheet.title = "Spot Check"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:N{len(rows) + 1}"

    visible_headers = [
        (1, "Company"),
        (2, "Year"),
        (3, "Sales"),
        (4, "Net Profit"),
        (5, "Equity Capital"),
        (6, "Reserves"),
        (7, "Database ROE"),
        (8, "Manual ROE"),
        (9, "Difference (%)"),
        (10, "Database Revenue CAGR"),
        (12, "Manual Revenue CAGR"),
        (13, "Difference (%)"),
        (14, "Status"),
    ]

    for column_index, header in visible_headers:
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    helper_header = worksheet.cell(row=1, column=11, value="Historical Sales (5Y)")
    helper_header.fill = HEADER_FILL
    helper_header.font = HEADER_FONT
    helper_header.alignment = CENTER
    worksheet.column_dimensions["K"].hidden = True

    for row_index, item in enumerate(rows, start=2):
        worksheet.cell(row=row_index, column=1, value=item["company_id"]).alignment = LEFT
        worksheet.cell(row=row_index, column=2, value=item["year"]).alignment = LEFT
        worksheet.cell(row=row_index, column=3, value=item["sales"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=4, value=item["net_profit"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=5, value=item["equity_capital"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=6, value=item["reserves"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=7, value=item["database_roe"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=8, value=f"=IF(OR(E{row_index}+F{row_index}=0,D{row_index}=\"\"),\"\",ROUND(D{row_index}/(E{row_index}+F{row_index})*100,2))").alignment = RIGHT
        worksheet.cell(row=row_index, column=9, value=f"=IF(OR(G{row_index}=\"\",H{row_index}=\"\"),\"\",ABS(G{row_index}-H{row_index}))").alignment = RIGHT
        worksheet.cell(row=row_index, column=10, value=item["database_revenue_cagr"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=11, value=item["historical_sales"]).alignment = RIGHT
        worksheet.cell(row=row_index, column=12, value=f"=IF(OR(C{row_index}<=0,K{row_index}<=0),\"\",ROUND(((C{row_index}/K{row_index})^(1/5)-1)*100,2))").alignment = RIGHT
        worksheet.cell(row=row_index, column=13, value=f"=IF(OR(J{row_index}=\"\",L{row_index}=\"\"),\"\",ABS(J{row_index}-L{row_index}))").alignment = RIGHT

        status = "PASS"
        if item["roe_difference"] >= 0.1 or item["revenue_cagr_difference"] >= 0.1:
            status = "FAIL"

        status_cell = worksheet.cell(row=row_index, column=14, value=status)
        status_cell.alignment = CENTER
        status_cell.font = BOLD_FONT
        status_cell.fill = PASS_FILL if status == "PASS" else FAIL_FILL

    for column_letter, width in {
        "A": 14,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 16,
        "F": 12,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 18,
        "K": 18,
        "L": 18,
        "M": 18,
        "N": 12,
    }.items():
        worksheet.column_dimensions[column_letter].width = width

    for row in worksheet.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=3, max_col=13):
        for cell in row:
            cell.number_format = "0.00"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def print_summary(row: dict[str, Any]) -> None:
    """Print the manual spot-check summary for a selected company-year."""

    print("===== SPOT CHECK =====")
    print()
    print(f"Company: {row['company_id']}")
    print(f"Year: {row['year']}")
    print(f"Database ROE: {row['database_roe']:.2f}")
    print(f"Manual ROE: {row['manual_roe']:.2f}")
    print(f"Difference: {row['roe_difference']:.4f}")
    print()
    print(f"Database Revenue CAGR: {row['database_revenue_cagr']:.2f}")
    print(f"Manual Revenue CAGR: {row['manual_revenue_cagr']:.2f}")
    print(f"Difference: {row['revenue_cagr_difference']:.4f}")
    print()
    print("PASS" if row["roe_difference"] < 0.1 and row["revenue_cagr_difference"] < 0.1 else "FAIL")
    print()


def print_sql_queries() -> None:
    """Print manual SQL verification queries for the selected companies."""

    queries = [
        "SELECT company_id, year, return_on_equity_pct, revenue_cagr_5yr\nFROM financial_ratios\nWHERE company_id='ABB';",
        "SELECT company_id, year, return_on_equity_pct, revenue_cagr_5yr\nFROM financial_ratios\nWHERE company_id='TCS';",
        "SELECT company_id, year, return_on_equity_pct, revenue_cagr_5yr\nFROM financial_ratios\nWHERE company_id='RELIANCE';",
    ]

    for query in queries:
        print(query)
        print()


def main() -> None:
    """Create the spot-check workbook and print validation details."""

    with connect_db(DB_PATH) as connection:
        candidate_rows = fetch_candidate_rows(connection)
        sales_history_rows = connection.execute(
            "SELECT company_id, year, sales FROM profitandloss WHERE company_id IS NOT NULL AND year IS NOT NULL"
        ).fetchall()

    sales_history = build_history_map(sales_history_rows)
    selected_rows = choose_spot_check_rows(candidate_rows, sales_history)

    if len(selected_rows) < 3:
        raise RuntimeError("Unable to find 3 valid companies for spot check")

    create_workbook(selected_rows, OUTPUT_PATH)

    for row in selected_rows:
        print_summary(row)

    print_sql_queries()
    print(f"Spot check workbook written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
